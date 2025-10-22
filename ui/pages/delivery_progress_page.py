# app/ui/pages/delivery_progress_page.py
import streamlit as st
import pandas as pd
from datetime import date, timedelta, datetime
from typing import Dict, Optional, Any
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class DeliveryProgressPage:
    """納入進度管理ページ"""
    
    def __init__(self, transport_service, auth_service=None):
        self.service = transport_service
        self.auth_service = auth_service

    def _can_edit_page(self) -> bool:
        """ページ編集権限チェック"""
        if not self.auth_service:
            return True
        user = st.session_state.get('user')
        if not user:
            return False
        return self.auth_service.can_edit_page(user['id'], "納入進度")
    
    def show(self):
        """ページ表示"""
        st.title("📋 納入進度管理")
        st.write("受注から出荷までの進捗を管理します。")

        # 権限チェック
        can_edit = self._can_edit_page()
        if not can_edit:
            st.warning("⚠️ この画面の編集権限がありません。閲覧のみ可能です。")

        tab1, tab2, tab3, tab4, tab5 = st.tabs([
            "📊 進度一覧",
            "✅ 実績登録",
            "➕ 新規登録",
            "📦 出荷実績",
            "🏭 社内注文"
        ])

        with tab1:
            self._show_progress_list(can_edit)
        with tab2:
            self._show_actual_registration(can_edit)
        with tab3:
            self._show_progress_registration(can_edit)
        with tab4:
            self._show_shipment_records()
        with tab5:
            self._show_internal_orders()
    
    def _show_progress_list(self, can_edit):
        """進度一覧表示"""
        st.header("📊 納入進度一覧")
        
        # サマリー表示
        try:
            summary = self.service.get_progress_summary()
            
            col1, col2, col3, col4, col5 = st.columns(5)
            
            with col1:
                st.metric("総オーダー数", summary.get('total_orders', 0))
            with col2:
                st.metric("未出荷", summary.get('unshipped', 0))
            with col3:
                st.metric("一部出荷", summary.get('partial', 0))
            with col4:
                st.metric("遅延", summary.get('delayed', 0), delta_color="inverse")
            with col5:
                st.metric("緊急", summary.get('urgent', 0), delta_color="inverse")
        
        except Exception as e:
            st.warning(f"サマリー取得エラー: {e}")

        # フィルター - デフォルトを過去10日間に変更
        st.subheader("🔍 フィルター")
        col_f1, col_f2, col_f3, col_f4 = st.columns(4)
        
        with col_f1:
            start_date = st.date_input(
                "納期（開始）",
                value=date.today() - timedelta(days=1),
                key="progress_start_date"
            )
        
        with col_f2:
            end_date = st.date_input(
                "納期（終了）",
                value=date.today()+timedelta(days=10),
                key="progress_end_date"
            )
        
        with col_f3:
            status_filter = st.multiselect(
                "ステータス",
                options=['未出荷', '計画済', '一部出荷', '出荷完了'],
                default=['未出荷', '計画済', '一部出荷', '出荷完了'],
                key="progress_status_filter"
            )

        with col_f4:
            product_filter = st.text_input(
                "製品コード（部分一致）",
                key="progress_product_filter"
            ).strip()
        # 進度データ取得
        try:
            progress_df = self.service.get_delivery_progress(start_date, end_date)

            with st.expander("計画進度の再計算"):
                # 製品リストを取得
                try:
                    products = self.service.product_repo.get_all_products()
                    if not products.empty:
                        product_options = {
                            f"{row['product_code']} - {row['product_name']}": row['id']
                            for _, row in products.iterrows()
                        }
                        selected_product = st.selectbox(
                            "製品コード",
                            options=list(product_options.keys()),
                            key="recalc_product_select"
                        )
                        product_id = product_options[selected_product]
                    else:
                        st.warning("製品が登録されていません")
                        product_id = None
                except:
                    st.error("製品データ取得エラー")
                    product_id = None

                recal_start_date = st.date_input("再計算開始日")
                recal_end_date = st.date_input("再計算終了日")

                col_recalc_single, col_recalc_all = st.columns(2)

                with col_recalc_single:
                    if st.button("選択製品のみ再計算", disabled=not can_edit):
                        if product_id:
                            self.service.recompute_planned_progress(product_id, recal_start_date, recal_end_date)
                            st.success("再計算が完了しました")
                        else:
                            st.error("製品を選択してください")

                with col_recalc_all:
                    if st.button("全製品を再計算", disabled=not can_edit):
                        self.service.recompute_planned_progress_all(recal_start_date, recal_end_date)
                        st.success("全ての製品に対する再計算が完了しました")

            # ▼ ここから追加：実績進度（shipped_remaining_quantity）の再計算
            with st.expander("実績進度の再計算（shipped_remaining_quantity）"):
                # 製品リストを取得
                try:
                    products = self.service.product_repo.get_all_products()
                    if not products.empty:
                        sr_product_options = {
                            f"{row['product_code']} - {row['product_name']}": row['id']
                            for _, row in products.iterrows()
                        }
                        sr_selected_product = st.selectbox(
                            "製品コード",
                            options=list(sr_product_options.keys()),
                            key="sr_product_select"
                        )
                        sr_product_id = sr_product_options[sr_selected_product]
                    else:
                        st.warning("製品が登録されていません")
                        sr_product_id = None
                except:
                    st.error("製品データ取得エラー")
                    sr_product_id = None

                sr_start_date = st.date_input("再計算開始日（実績）", key="sr_start_date")
                sr_end_date = st.date_input("再計算終了日（実績）", key="sr_end_date")

                col_sr_one, col_sr_all = st.columns(2)

                with col_sr_one:
                    if st.button("選択製品の実績進度を再計算", key="btn_sr_one", disabled=not can_edit):
                        if sr_product_id:
                            self.service.recompute_shipped_remaining(sr_product_id, sr_start_date, sr_end_date)
                            st.success("実績進度の再計算が完了しました")
                        else:
                            st.error("製品を選択してください")

                with col_sr_all:
                    if st.button("全製品の実績進度を再計算", key="btn_sr_all", disabled=not can_edit):
                        self.service.recompute_shipped_remaining_all(sr_start_date, sr_end_date)
                        st.success("全製品の実績進度の再計算が完了しました")
                              
            if not progress_df.empty:
                # ステータスフィルター適用
                if status_filter:
                    progress_df = progress_df[progress_df['status'].isin(status_filter)]
                if product_filter:
                    progress_df = progress_df[
                        progress_df['product_code'].fillna('').str.contains(product_filter, case=False, na=False)
                    ]
                
                # 表示形式選択を追加
                st.subheader("📋 表示形式")
                view_mode = st.radio(
                    "表示モード",
                    options=['一覧表示', 'マトリックス表示（日付×製品）'],
                    horizontal=True,
                    key="view_mode_selector"
                )
                
                if view_mode == 'マトリックス表示（日付×製品）':
                    self._show_matrix_view(progress_df, can_edit)
                else:
                    # 既存の一覧表示
                    # 緊急度フラグ追加
                    progress_df['days_to_delivery'] = (
                        pd.to_datetime(progress_df['delivery_date']) - pd.Timestamp(date.today())
                    ).dt.days
                    
                    progress_df['urgency'] = progress_df.apply(
                        lambda row: '🔴遅延' if row['days_to_delivery'] < 0 and row['status'] != '出荷完了'
                        else '🟡緊急' if 0 <= row['days_to_delivery'] <= 3 and row['status'] != '出荷完了'
                        else '🟢',
                        axis=1
                    )
                    
                    # 計画進度と進度を計算
                    progress_df['planned_progress'] = (
                        progress_df.get('planned_quantity', 0).fillna(0) -
                        progress_df.get('order_quantity', 0).fillna(0)
                    )
                    progress_df['actual_progress'] = (
                        progress_df.get('shipped_quantity', 0).fillna(0) -
                        progress_df.get('order_quantity', 0).fillna(0)
                    )

                    st.subheader("🖊️ 手動計画数量の一括編集")
                    # 上の一覧表示と同じ列構成にする（IDなし）
                    editor_columns = ['urgency', 'product_code', 'product_name', 'delivery_date', 'order_quantity']

                    if 'manual_planning_quantity' in progress_df.columns:
                        editor_columns.append('manual_planning_quantity')
                    if 'planned_quantity' in progress_df.columns:
                        editor_columns.append('planned_quantity')

                    editor_columns.extend(['planned_progress', 'shipped_quantity', 'actual_progress', 'remaining_quantity', 'status'])

                    # IDは保存処理のために別途保持
                    editor_source = progress_df[editor_columns].copy()
                    editor_source.insert(0, 'id', progress_df['id'])
                    editor_source = editor_source.reset_index(drop=True)

                    original_editor = editor_source.copy()

                    # 手動計画のみFloat64型に変換（編集可能にするため）
                    if 'manual_planning_quantity' in editor_source.columns:
                        editor_source['manual_planning_quantity'] = editor_source['manual_planning_quantity'].astype('Float64')

                    # ID列を非表示にして12列で表示
                    edited_table = st.data_editor(
                        editor_source,
                        num_rows="fixed",
                        hide_index=True,
                        use_container_width=True,
                        column_config={
                            'id': None,  # ID列を非表示
                            'urgency': st.column_config.TextColumn('緊急度'),
                            'product_code': st.column_config.TextColumn('製品コード'),
                            'product_name': st.column_config.TextColumn('製品名'),
                            'delivery_date': st.column_config.DateColumn('納期', format='YYYY-MM-DD'),
                            'order_quantity': st.column_config.NumberColumn('受注数', format='%d'),
                            'manual_planning_quantity': st.column_config.NumberColumn('手動計画', min_value=0, step=1),
                            'planned_quantity': st.column_config.NumberColumn('計画数', format='%d'),
                            'planned_progress': st.column_config.NumberColumn('計画進度', format='%d'),
                            'shipped_quantity': st.column_config.NumberColumn('出荷済', format='%d'),
                            'actual_progress': st.column_config.NumberColumn('進度', format='%d'),
                            'remaining_quantity': st.column_config.NumberColumn('残数', format='%d'),
                            'status': st.column_config.TextColumn('ステータス'),
                        },
                        disabled=['urgency', 'product_code', 'product_name', 'delivery_date', 'order_quantity', 'planned_quantity', 'planned_progress', 'shipped_quantity', 'actual_progress', 'remaining_quantity', 'status'],
                        key="manual_plan_editor",
                    )
                    st.markdown(":red[手動計画列のみ編集できます。注意：１増減分ではなく、変更後の数値を入力。２　キャンセルするとき、０入力ではなくNoneなるように消す]")

                    if st.button("手動計画を保存", type="primary", key="save_manual_plans", disabled=not can_edit):
                        updated_count = 0
                        for idx, row in edited_table.iterrows():
                            new_val = row['manual_planning_quantity']
                            orig_val = original_editor.loc[idx, 'manual_planning_quantity']
                            if pd.isna(new_val) or new_val == '':
                                new_db_val = None
                            else:
                                try:
                                    new_db_val = int(new_val)
                                except (TypeError, ValueError):
                                    st.warning(f"ID {int(row['id'])} の値が無効です。")
                                    continue

                            if pd.isna(orig_val):
                                orig_compare = None
                            else:
                                try:
                                    orig_compare = int(orig_val)
                                except (TypeError, ValueError):
                                    orig_compare = None

                            if orig_compare == new_db_val:
                                continue

                            success = self.service.update_delivery_progress(
                                int(row['id']),
                                {'manual_planning_quantity': new_db_val}
                            )
                            if success:
                                updated_count += 1
                            else:
                                st.error(f"ID {int(row['id'])} の更新に失敗しました。")

                        if updated_count:
                            st.success(f"{updated_count} 件の手動計画を更新しました。")
                            st.rerun()
                        else:
                            st.info("変更はありませんでした。")
                    
                    # 詳細編集・出荷実績入力
                    st.subheader("📝 詳細編集・出荷実績入力")
                    
                    if not progress_df.empty:
                        # オーダー選択 - 製品コード表示
                        order_options = {
                            f"{row['order_id']} - {row['product_code']} ({row['delivery_date']})": row['id']
                            for _, row in progress_df.iterrows()
                        }
                        
                        selected_order_key = st.selectbox(
                            "編集するオーダーを選択",
                            options=list(order_options.keys()),
                            key="progress_edit_selector"
                        )
                        
                        if selected_order_key:
                            progress_id = order_options[selected_order_key]
                            progress_row = progress_df[progress_df['id'] == progress_id].iloc[0]
                            
                            # タブで編集と出荷実績を分離
                            edit_tab, shipment_tab = st.tabs(["📝 進度編集", "📦 出荷実績入力"])
                            
                            with edit_tab:
                                with st.form(f"edit_progress_{progress_id}"):
                                    st.write("**進度情報を編集**")
                                    
                                    col_e1, col_e2 = st.columns(2)
                                    
                                    with col_e1:
                                        new_delivery_date = st.date_input(
                                            "納期",
                                            value=progress_row['delivery_date'],
                                            key=f"delivery_{progress_id}"
                                        )
                                        new_priority = st.number_input(
                                            "優先度（1-10）",
                                            min_value=1,
                                            max_value=10,
                                            value=int(progress_row.get('priority', 5)),
                                            key=f"priority_{progress_id}"
                                        )
                                    
                                    with col_e2:
                                        new_status = st.selectbox(
                                            "ステータス",
                                            options=['未出荷', '計画済', '一部出荷', '出荷完了', 'キャンセル'],
                                            index=['未出荷', '計画済', '一部出荷', '出荷完了', 'キャンセル'].index(progress_row['status']) if progress_row['status'] in ['未出荷', '計画済', '一部出荷', '出荷完了', 'キャンセル'] else 0,
                                            key=f"status_{progress_id}"
                                        )
                                        new_notes = st.text_area(
                                            "備考",
                                            value=progress_row.get('notes', '') or '',
                                            key=f"notes_{progress_id}"
                                        )
                                    
                                    manual_value = progress_row.get('manual_planning_quantity')
                                    use_manual = st.checkbox(
                                        "手動計画数量を指定",
                                        value=pd.notna(manual_value),
                                        key=f"use_manual_{progress_id}"
                                    )
                                    if pd.notna(manual_value):
                                        manual_default = int(manual_value)
                                    else:
                                        manual_default = int(progress_row.get('order_quantity', 0) or 0)
                                    manual_quantity = st.number_input(
                                        "手動計画数量",
                                        min_value=0,
                                        value=manual_default,
                                        step=1,
                                        key=f"manual_qty_{progress_id}",
                                        disabled=not use_manual
                                    )
                                    
                                    submitted = st.form_submit_button("💾 更新", type="primary", disabled=not can_edit)
                                    
                                    if submitted:
                                        update_data = {
                                            'delivery_date': new_delivery_date,
                                            'priority': new_priority,
                                            'status': new_status,
                                            'notes': new_notes,
                                            'manual_planning_quantity': int(manual_quantity) if use_manual else None
                                        }
                                        
                                        success = self.service.update_delivery_progress(progress_id, update_data)
                                        if success:
                                            st.success("進度を更新しました")
                                            st.rerun()
                                        else:
                                            st.error("進度更新に失敗しました")
                            
                            # 出荷実績入力タブ
                            with shipment_tab:
                                # 現在の出荷状況を表示
                                manual_display = progress_row.get('manual_planning_quantity')
                                manual_display = int(manual_display) if pd.notna(manual_display) else '未設定'
                                st.info(f"""
                                **現在の状況:**
                                - 受注数: {progress_row.get('order_quantity', 0)}
                                - 計画数: {progress_row.get('planned_quantity', 0)}
                                - 手動計画: {manual_display}
                                - 出荷済: {progress_row.get('shipped_quantity', 0)}
                                - 残数: {progress_row.get('remaining_quantity', 0)}
                                """)
                                
                                with st.form(f"shipment_form_{progress_id}"):
                                    st.write("**出荷実績を入力**")
                                    
                                    col_s1, col_s2 = st.columns(2)
                                    
                                    with col_s1:
                                        shipment_date = st.date_input(
                                            "出荷日 *",
                                            value=date.today(),
                                            key=f"ship_date_{progress_id}"
                                        )
                                        
                                        # トラック選択
                                        try:
                                            trucks_df = self.service.get_trucks()
                                            
                                            if not trucks_df.empty:
                                                truck_options = dict(zip(trucks_df['name'], trucks_df['id']))
                                                selected_truck = st.selectbox(
                                                    "使用トラック *",
                                                    options=list(truck_options.keys()),
                                                    key=f"ship_truck_{progress_id}"
                                                )
                                                truck_id = truck_options[selected_truck]
                                            else:
                                                st.warning("トラックが登録されていません")
                                                truck_id = None
                                        except:
                                            st.warning("トラック情報の取得に失敗しました")
                                            truck_id = None
                                        
                                        remaining_qty = int(progress_row.get('remaining_quantity', 0))
                                        if remaining_qty > 0:
                                            shipped_quantity = st.number_input(
                                                "出荷数量 *",
                                                min_value=1,
                                                max_value=remaining_qty,
                                                value=min(100, remaining_qty),
                                                key=f"ship_qty_{progress_id}"
                                            )
                                        else:
                                            st.warning("出荷可能な数量がありません")
                                            shipped_quantity = 0
                                    # delivery_progress_page.py の該当箇所を修正
                                    with col_s2:
                                        driver_name = st.text_input(
                                            "ドライバー名",
                                            key=f"driver_{progress_id}"
                                        )
                                        
                                        # トラックのデフォルト時刻を取得
                                        default_dep_time = None
                                        default_arr_time = None
                                        
                                        if truck_id and not trucks_df.empty:
                                            try:
                                                truck_row = trucks_df[trucks_df['id'] == truck_id]
                                                if not truck_row.empty:
                                                    truck_info = truck_row.iloc[0]
                                                    # departure_time と arrival_time カラムを使用
                                                    if 'departure_time' in truck_info and pd.notna(truck_info['departure_time']):
                                                        default_dep_time = truck_info['departure_time']
                                                    if 'arrival_time' in truck_info and pd.notna(truck_info['arrival_time']):
                                                        default_arr_time = truck_info['arrival_time']
                                            except Exception as e:
                                                print(f"トラック時刻取得エラー: {e}")
                                        
                                        # デフォルト値を設定(トラック設定時刻がなければNone)
                                        actual_departure = st.time_input(
                                            "実出発時刻",
                                            value=default_dep_time,
                                            key=f"dep_time_{progress_id}"
                                        )
                                        
                                        actual_arrival = st.time_input(
                                            "実到着時刻",
                                            value=default_arr_time,
                                            key=f"arr_time_{progress_id}"
                                        )
                                        
                                        shipment_notes = st.text_area(
                                            "備考",
                                            key=f"ship_notes_{progress_id}"
                                        )

                                    # 出荷実績登録ボタン
                                    
                                    ship_submitted = st.form_submit_button("📦 出荷実績を登録", type="primary", disabled=not can_edit)
                                    
                                    if ship_submitted:
                                        if not truck_id:
                                            st.error("トラックを選択してください")
                                        elif shipped_quantity <= 0:
                                            st.error("出荷数量を入力してください")
                                        else:
                                            shipment_data = {
                                                'progress_id': progress_id,
                                                'truck_id': truck_id,
                                                'shipment_date': shipment_date,
                                                'shipped_quantity': shipped_quantity,
                                                'driver_name': driver_name,
                                                'actual_departure_time': actual_departure,
                                                'actual_arrival_time': actual_arrival,
                                                'notes': shipment_notes
                                            }
                                            
                                            success = self.service.create_shipment_record(shipment_data)
                                            if success:
                                                st.success(f"✅ 出荷実績を登録しました（{shipped_quantity}個）")
                                                st.balloons()
                                                st.rerun()
                                            else:
                                                st.error("❌ 出荷実績登録に失敗しました")
                            
                            # 削除ボタンは外に配置
                            st.markdown("---")
                            col_del1, col_del2 = st.columns([1, 5])
                            with col_del1:
                                if st.button(f"🗑️ 削除", key=f"delete_progress_{progress_id}", type="secondary", disabled=not can_edit):
                                    success = self.service.delete_delivery_progress(progress_id)
                                    if success:
                                        st.success("進度を削除しました")
                                        st.rerun()
                                    else:
                                        st.error("進度削除に失敗しました")
            
            else:
                st.info("指定期間内に納入進度データがありません")
        
        except Exception as e:
            st.error(f"進度一覧エラー: {e}")
    
    def _show_matrix_view(self, progress_df: pd.DataFrame, can_edit):
        """マトリックス表示（横軸=日付、縦軸=製品コード×状態）- 編集可能"""
        
        # 製品名マッピング作成
        product_names = progress_df.groupby('product_code')['product_name'].first().to_dict()
        
        # 製品コード一覧を取得
        product_codes = sorted(progress_df['product_code'].unique())
        
        # 日付一覧を取得（文字列形式）
        dates = sorted(progress_df['delivery_date'].unique())
        date_columns = [d.strftime('%m月%d日') for d in dates]
        
        st.write(f"**製品数**: {len(product_codes)}")
        st.write(f"**日付数**: {len(dates)}")
        
        # オーダーIDマッピング（更新用）
        order_mapping = {}  # {(product_code, date_str): order_id}
        for _, row in progress_df.iterrows():
            key = (row['product_code'], row['delivery_date'].strftime('%m月%d日'))
            order_mapping[key] = row['id']
        
        # 結果を格納するリスト
        result_rows = []
        
        for product_code in product_codes:
            product_data = progress_df[progress_df['product_code'] == product_code]
            
            # 各指標の行を作成
            order_row = {'製品コード': product_code, '状態': '受注数', 'row_type': 'order'}
            planned_row = {'製品コード': '', '状態': '納入計画数', 'row_type': 'planned'}
            planned_progress_row = {'製品コード': '', '状態': '計画進度', 'row_type': 'planned_progress'}
            shipped_row = {'製品コード': '', '状態': '納入実績', 'row_type': 'shipped'}
            progress_row = {'製品コード': '', '状態': '進度', 'row_type': 'progress'}
            keisen_row = {'製品コード': '', '状態': '___', 'row_type': 'ーーー'}
            
            cumulative_order = 0
            cumulative_planned = 0
            cumulative_shipped = 0
            
            for idx, (date_obj, date_str) in enumerate(zip(dates, date_columns)):
                # その日のデータを取得
                day_data = product_data[product_data['delivery_date'] == date_obj]
                
                if not day_data.empty:
                    row = day_data.iloc[0]
                    
                    order_qty = int(row['order_quantity']) if pd.notna(row['order_quantity']) else 0
                    
                    # planned_quantity の安全な取得
                    if 'planned_quantity' in day_data.columns and pd.notna(row['planned_quantity']):
                        planned_qty = int(row['planned_quantity'])
                    else:
                        planned_qty = 0
                    
                    # shipped_quantity の安全な取得
                    if 'shipped_quantity' in day_data.columns and pd.notna(row['shipped_quantity']):
                        shipped_qty = int(row['shipped_quantity'])
                    else:
                        shipped_qty = 0
                    
                    cumulative_order += order_qty
                    cumulative_planned += planned_qty
                    cumulative_shipped += shipped_qty
                    
                    order_row[date_str] = order_qty
                    planned_row[date_str] = planned_qty
                    planned_progress_row[date_str] = cumulative_planned - cumulative_order
                    shipped_row[date_str] = shipped_qty
                else:
                    order_row[date_str] = 0
                    planned_row[date_str] = 0
                    planned_progress_row[date_str] = cumulative_planned - cumulative_order
                    shipped_row[date_str] = 0
                
                # 進度 = 累計出荷 - 累計受注
                progress = cumulative_shipped - cumulative_order
                progress_row[date_str] = int(progress)
            
            result_rows.extend([order_row, planned_row, planned_progress_row, shipped_row, progress_row, keisen_row])
        
        # DataFrameに変換
        result_df = pd.DataFrame(result_rows)
        
        # カラムの順序を整理
        columns = ['製品コード', '状態', 'row_type'] + date_columns
        result_df = result_df[columns]
        
        st.write("---")
        st.write("**日付×製品マトリックス（受注・計画・実績・進度）**")
        
        # 修正: 列を固定表示（製品コードと状態列を固定）
        edited_df = st.data_editor(
            result_df,
            use_container_width=True,
            hide_index=True,
            disabled=['製品コード', '状態', 'row_type'],  # 編集不可カラム
            column_config={
                "製品コード": st.column_config.TextColumn(
                    "製品コード", 
                    width="medium",
                    pinned=True
                ),
                "状態": st.column_config.TextColumn(
                    "状態", 
                    width="small",
                    pinned=True
                ),
                "row_type": None,  # 非表示
                **{col: st.column_config.NumberColumn(col, step=1) for col in date_columns}
            },
            key="matrix_editor"
        )
        
        # 保存ボタン
        col_save1, col_save2 = st.columns([1, 5])
        
        with col_save1:
            if st.button("💾 変更を保存", type="primary", use_container_width=True, disabled=not can_edit):
                # 変更を検出して保存
                changes_saved = self._save_matrix_changes(
                    original_df=result_df,
                    edited_df=edited_df,
                    order_mapping=order_mapping,
                    product_codes=product_codes,
                    dates=dates,
                    date_columns=date_columns,
                    progress_df=progress_df
                )
                
                if changes_saved:
                    st.success("✅ 変更を保存しました")
                    st.rerun()
                else:
                    st.info("変更はありませんでした")
        
        with col_save2:
            st.caption("※ 「計画進度」「進度」行は自動計算されます（計画進度=累計計画 - 累計受注、進度=累計出荷 - 累計受注）")
        
        # 説明
        with st.expander("📋 表の見方"):
            st.write("""
            **各行の意味:**
            - **受注数**: その日の受注数量（編集不可）
            - **納入計画数**: 積載計画で設定された数量（編集可）
            - **計画進度**: 累計計画 - 累計受注（自動計算）
            - **納入実績**: 実際に出荷した数量（編集可）
            - **進度**: 累計出荷 - 累計受注（自動計算、マイナスは未納分）
            
            **編集方法:**
            1. 「納入計画数」または「納入実績」のセルをダブルクリック
            2. 数値を入力
            3. 「💾 変更を保存」ボタンをクリック
            """)

    def _save_matrix_changes(self, original_df, edited_df, order_mapping, 
                            product_codes, dates, date_columns, progress_df):
        """マトリックスの変更をデータベースに保存"""
        
        changes_made = False
        
        for product_code in product_codes:
            for date_obj, date_str in zip(dates, date_columns):
                # オーダーIDを取得
                order_key = (product_code, date_str)
                if order_key not in order_mapping:
                    continue
                
                order_id = order_mapping[order_key]
                
                # 元データを取得
                original_data = progress_df[
                    (progress_df['product_code'] == product_code) & 
                    (progress_df['delivery_date'] == date_obj)
                ]
                
                if original_data.empty:
                    continue
                
                # NaN対応
                original_planned = int(original_data['planned_quantity'].iloc[0]) if pd.notna(original_data['planned_quantity'].iloc[0]) else 0
                original_shipped = int(original_data['shipped_quantity'].iloc[0]) if pd.notna(original_data['shipped_quantity'].iloc[0]) else 0
                
                # 編集後のデータを取得
                planned_rows = edited_df[
                    (edited_df['row_type'] == 'planned') &
                    ((edited_df['製品コード'] == product_code) | (edited_df['製品コード'] == ''))
                ]
                
                shipped_rows = edited_df[
                    (edited_df['row_type'] == 'shipped') &
                    ((edited_df['製品コード'] == product_code) | (edited_df['製品コード'] == ''))
                ]
                
                # 納入計画数の変更チェック
                if not planned_rows.empty and date_str in planned_rows.columns:
                    product_planned_rows = planned_rows[
                        (planned_rows.index > edited_df[edited_df['製品コード'] == product_code].index.min()) &
                        (planned_rows.index < edited_df[edited_df['製品コード'] == product_code].index.min() + 4)
                    ]
                    
                    if not product_planned_rows.empty:
                        new_planned = int(product_planned_rows.iloc[0][date_str]) if pd.notna(product_planned_rows.iloc[0][date_str]) else 0
                        
                        if new_planned != original_planned:
                            update_data = {'planned_quantity': new_planned}
                            success = self.service.update_delivery_progress(order_id, update_data)
                            if success:
                                changes_made = True
                                print(f"✅ 計画数更新: order_id={order_id}, {original_planned} → {new_planned}")
                
                # 納入実績の変更チェック
                if not shipped_rows.empty and date_str in shipped_rows.columns:
                    product_shipped_rows = shipped_rows[
                        (shipped_rows.index > edited_df[edited_df['製品コード'] == product_code].index.min()) &
                        (shipped_rows.index < edited_df[edited_df['製品コード'] == product_code].index.min() + 4)
                    ]
                    
                    if not product_shipped_rows.empty:
                        new_shipped = int(product_shipped_rows.iloc[0][date_str]) if pd.notna(product_shipped_rows.iloc[0][date_str]) else 0
                        
                        # ✅ 修正: 直接 delivery_progress を更新
                        if new_shipped != original_shipped:
                            # 1. delivery_progress.shipped_quantity を直接更新
                            update_data = {'shipped_quantity': new_shipped}
                            success = self.service.update_delivery_progress(order_id, update_data)
                            
                            if success:
                                changes_made = True
                                print(f"✅ 実績更新: order_id={order_id}, {original_shipped} → {new_shipped}")
                                
                                # 2. 差分があれば出荷実績レコードも作成（履歴として）
                                diff = new_shipped - original_shipped
                                if diff > 0:
                                    shipment_data = {
                                        'progress_id': order_id,
                                        'truck_id': 1,
                                        'shipment_date': date_obj,
                                        'shipped_quantity': diff,
                                        'driver_name': 'マトリックス入力',
                                        'actual_departure_time': None,
                                        'actual_arrival_time': None,
                                        'notes': f'マトリックスから直接入力（累計: {new_shipped}）'
                                    }
                                    self.service.create_shipment_record(shipment_data)
        
        return changes_made

    def _show_progress_registration(self, can_edit):
        """新規登録"""
        st.header("➕ 新規納入進度登録")

        if not can_edit:
            st.info("編集権限がないため、新規登録はできません")
            return
        
        with st.form("create_progress_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**オーダー情報**")
                order_id = st.text_input("オーダーID *", placeholder="例: ORD-2025-001")
                
                # 製品選択
                try:
                    products = self.service.product_repo.get_all_products()
                    if not products.empty:
                        product_options = {
                            f"{row['product_code']} - {row['product_name']}": row['id']
                            for _, row in products.iterrows()
                        }
                        selected_product = st.selectbox("製品 *", options=list(product_options.keys()))
                        product_id = product_options[selected_product]
                    else:
                        st.warning("製品が登録されていません")
                        product_id = None
                except:
                    st.error("製品データ取得エラー")
                    product_id = None
                
                order_date = st.date_input("受注日 *", value=date.today())
                delivery_date = st.date_input("納期 *", value=date.today() + timedelta(days=7))
                order_quantity = st.number_input("受注数量 *", min_value=1, value=100, step=1)
            
            with col2:
                st.write("**得意先情報**")
                customer_code = st.text_input("得意先コード", placeholder="例: C001")
                customer_name = st.text_input("得意先名", placeholder="例: 株式会社〇〇")
                delivery_location = st.text_input("納入先", placeholder="例: 東京工場")
                priority = st.number_input("優先度（1-10）", min_value=1, max_value=10, value=5)
                notes = st.text_area("備考")
            
            submitted = st.form_submit_button("➕ 登録", type="primary")
            
            if submitted:
                if not order_id or not product_id:
                    st.error("オーダーIDと製品は必須です")
                else:
                    progress_data = {
                        'order_id': order_id,
                        'product_id': product_id,
                        'order_date': order_date,
                        'delivery_date': delivery_date,
                        'order_quantity': order_quantity,
                        'customer_code': customer_code,
                        'customer_name': customer_name,
                        'delivery_location': delivery_location,
                        'priority': priority,
                        'notes': notes
                    }
                    
                    progress_id = self.service.create_delivery_progress(progress_data)
                    if progress_id > 0:
                        st.success(f"納入進度を登録しました（ID: {progress_id}）")
                        st.balloons()
                        st.rerun()
                    else:
                        st.error("納入進度登録に失敗しました")
    
    def _show_actual_registration(self, can_edit):
        """実績登録"""
        st.header("✅ 積込実績登録")

        if not can_edit:
            st.info("編集権限がないため、実績登録はできません")
            return
        
        try:
            trucks_df = self.service.get_trucks()
        except Exception as e:
            st.error(f"トラック情報の取得に失敗しました: {e}")
            return
        
        if trucks_df is None or trucks_df.empty:
            st.info("トラックマスタが空です。先にトラックを登録してください。")
            return
        
        truck_options = {
            str(row["name"]): int(row["id"])
            for _, row in trucks_df.iterrows()
            if pd.notna(row.get("name")) and pd.notna(row.get("id"))
        }
        
        if not truck_options:
            st.info("選択可能なトラックがありません。")
            return
        
        col1, col2 = st.columns(2)
        with col1:
            loading_date = st.date_input(
                "積込日",
                value=date.today(),
                key="actual_loading_date"
            )
        with col2:
            truck_name = st.selectbox(
                "トラック",
                options=list(truck_options.keys()),
                key="actual_truck_select"
            )
        
        selected_truck_id = truck_options.get(truck_name)
        if not selected_truck_id:
            st.warning("トラックを選択してください。")
            return
        
        try:
            plan_items = self.service.get_loading_plan_details_by_date(loading_date, selected_truck_id)
        except Exception as e:
            st.error(f"積載計画の取得に失敗しました: {e}")
            return
        
        if not plan_items:
            st.info("指定条件に該当する積載計画がありません。")
            return
        
        plan_df = pd.DataFrame(plan_items)
        if plan_df.empty or 'id' not in plan_df.columns:
            st.error("積載計画明細の形式が不正です。")
            return
        
        plan_df = plan_df.set_index('id')
        
        if 'delivery_date' in plan_df.columns:
            plan_df['delivery_date'] = pd.to_datetime(
                plan_df['delivery_date'], errors='coerce'
            ).dt.date
        plan_df['delivery_date'] = plan_df['delivery_date'].fillna(loading_date)
        
        if 'trip_number' in plan_df.columns:
            plan_df['trip_number'] = pd.to_numeric(plan_df['trip_number'], errors='coerce').fillna(1).astype(int)
        else:
            plan_df['trip_number'] = 1
        
        plan_df['num_containers'] = pd.to_numeric(plan_df.get('num_containers', 0), errors='coerce').fillna(0).astype(int)
        plan_df['total_quantity'] = pd.to_numeric(plan_df.get('total_quantity', 0), errors='coerce').fillna(0).astype(int)
        plan_df['planned_quantity'] = plan_df['total_quantity']
        
        progress_cache: Dict[int, Optional[Dict[str, Any]]] = {}
        missing_progress: list[str] = []
        
        plan_df['current_shipped'] = None
        plan_df['current_status'] = None
        
        for detail_id, row in plan_df.iterrows():
            product_id = row.get('product_id')
            try:
                product_id_int = int(product_id)
            except (TypeError, ValueError):
                progress_cache[detail_id] = None
                missing_progress.append(f"{row.get('product_code', '') or '不明'}")
                continue
            
            delivery_value = row.get('delivery_date') or loading_date
            if isinstance(delivery_value, pd.Timestamp):
                delivery_value = delivery_value.to_pydatetime().date()
            elif isinstance(delivery_value, datetime):
                delivery_value = delivery_value.date()
            elif isinstance(delivery_value, str):
                try:
                    delivery_value = datetime.strptime(delivery_value, "%Y-%m-%d").date()
                except ValueError:
                    delivery_value = loading_date
            
            plan_df.at[detail_id, 'delivery_date'] = delivery_value
            
            try:
                progress = self.service.get_delivery_progress_by_product_and_date(product_id_int, delivery_value)
            except Exception as e:
                st.warning(f"納入進度の取得に失敗しました（製品ID:{product_id_int}）: {e}")
                progress = None
            
            progress_cache[detail_id] = progress
            
            if progress:
                shipped_val = progress.get('shipped_quantity')
                plan_df.at[detail_id, 'current_shipped'] = int(shipped_val) if shipped_val is not None else 0
                plan_df.at[detail_id, 'current_status'] = progress.get('status')
            else:
                plan_df.at[detail_id, 'current_shipped'] = None
                plan_df.at[detail_id, 'current_status'] = None
                missing_progress.append(f"{row.get('product_code', '') or '不明'}（{delivery_value}）")
        
        product_codes = plan_df.get('product_code', pd.Series('', index=plan_df.index))
        product_names = plan_df.get('product_name', pd.Series('', index=plan_df.index))
        
        display_df = pd.DataFrame(
            {
                "積込順": plan_df['trip_number'],
                "製品コード": product_codes,
                "製品名": product_names,
                "納入日": plan_df['delivery_date'],
                "計画数量": plan_df['planned_quantity'],
                "既出荷数量": plan_df['current_shipped'].fillna(0).astype(int),
                "実績数量": plan_df['planned_quantity']
            },
            index=plan_df.index
        )
        display_df.index.name = "detail_id"
        
        st.caption("計画数量をベースに実績数量を入力してください。不要な行は0のままにします。")
        if missing_progress:
            st.warning("納入進度が見つからない明細があります: " + "、".join(sorted(set(missing_progress))))
        
        form_key = f"actual_registration_form_{selected_truck_id}_{loading_date.isoformat()}"
        with st.form(form_key):
            edited_df = st.data_editor(
                display_df,
                key=f"actual_editor_{selected_truck_id}_{loading_date.isoformat()}",
                use_container_width=True,
                hide_index=True,
                column_config={
                    "積込順": st.column_config.NumberColumn("積込順", disabled=True),
                    "製品コード": st.column_config.TextColumn("製品コード", disabled=True),
                    "製品名": st.column_config.TextColumn("製品名", disabled=True),
                    "納入日": st.column_config.DateColumn("納入日", disabled=True, format="YYYY-MM-DD"),
                    "計画数量": st.column_config.NumberColumn("計画数量", disabled=True),
                    "既出荷数量": st.column_config.NumberColumn("既出荷数量", disabled=True),
                    "実績数量": st.column_config.NumberColumn("実績数量", min_value=0, step=1)
                }
            )
            
            driver_name = st.text_input("ドライバー名", key=f"actual_driver_{selected_truck_id}")
            notes = st.text_area(
                "備考（必要に応じて入力）",
                key=f"actual_notes_{selected_truck_id}",
                placeholder=f"例: {truck_name} {loading_date} 積込"
            )
            
            submitted = st.form_submit_button("実績を登録", type="primary")
            
            if submitted:
                if edited_df.empty:
                    st.info("登録対象の明細がありません。")
                    return
                
                registered = 0
                failed_entries: list[str] = []
                missing_entries: list[str] = []
                
                for detail_id, row in edited_df.iterrows():
                    try:
                        detail_id_int = int(detail_id)
                    except (TypeError, ValueError):
                        continue
                    
                    quantity_value = pd.to_numeric(row.get("実績数量"), errors='coerce')
                    if pd.isna(quantity_value) or quantity_value <= 0:
                        continue
                    
                    progress = progress_cache.get(detail_id_int)
                    plan_row = plan_df.loc[detail_id_int]
                    
                    if not progress:
                        missing_entries.append(f"{plan_row.get('product_code', '') or '不明'}（{plan_row.get('delivery_date')}）")
                        continue
                    
                    shipment_data = {
                        'progress_id': progress['id'],
                        'truck_id': selected_truck_id,
                        'shipment_date': loading_date,
                        'shipped_quantity': int(quantity_value),
                        'container_id': plan_row.get('container_id'),
                        'num_containers': plan_row.get('num_containers'),
                        'driver_name': driver_name,
                        'notes': notes
                    }
                    
                    success = self.service.create_shipment_record(shipment_data)
                    if success:
                        registered += 1
                    else:
                        failed_entries.append(f"{plan_row.get('product_code', '') or '不明'}（{plan_row.get('delivery_date')}）")
                
                if registered:
                    st.success(f"{registered} 件の実績を登録しました。")
                    st.balloons()
                if failed_entries:
                    st.error("登録に失敗した明細: " + "、".join(failed_entries))
                if missing_entries:
                    st.warning("納入進度が見つからず登録できなかった明細: " + "、".join(missing_entries))
                
                if registered and not failed_entries:
                    st.info("他のタブで最新の実績を確認できます。")
                    st.rerun()
    
    def _show_shipment_records(self):
        """出荷実績表示"""
        st.header("📦 出荷実績一覧")
        
        # フィルター
        col_f1, col_f2 = st.columns(2)
        
        with col_f1:
            filter_start = st.date_input(
                "出荷日（開始）",
                value=date.today() - timedelta(days=7),
                key="shipment_start_filter"
            )
        
        with col_f2:
            filter_end = st.date_input(
                "出荷日（終了）",
                value=date.today(),
                key="shipment_end_filter"
            )
        
        try:
            shipment_df = self.service.get_shipment_records()
            
            if not shipment_df.empty:
                # 日付フィルター適用
                shipment_df['shipment_date'] = pd.to_datetime(shipment_df['shipment_date']).dt.date
                filtered_df = shipment_df[
                    (shipment_df['shipment_date'] >= filter_start) &
                    (shipment_df['shipment_date'] <= filter_end)
                ]
                
                if not filtered_df.empty:
                    # 表示用データフレーム
                    display_cols = ['shipment_date', 'order_id', 'product_code', 'product_name', 
                                  'truck_name', 'shipped_quantity', 'driver_name']
                    
                    # カラムが存在するかチェック
                    available_cols = [col for col in display_cols if col in filtered_df.columns]
                    
                    if 'num_containers' in filtered_df.columns:
                        available_cols.append('num_containers')
                    
                    display_df = filtered_df[available_cols].copy()
                    
                    # カラム名を日本語に
                    column_mapping = {
                        'shipment_date': '出荷日',
                        'order_id': 'オーダーID',
                        'product_code': '製品コード',
                        'product_name': '製品名',
                        'truck_name': 'トラック',
                        'shipped_quantity': '出荷数量',
                        'num_containers': '容器数',
                        'driver_name': 'ドライバー'
                    }
                    
                    display_df.columns = [column_mapping.get(col, col) for col in display_df.columns]
                    
                    st.dataframe(
                        display_df,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "出荷日": st.column_config.DateColumn("出荷日", format="YYYY-MM-DD"),
                        }
                    )
                    
                    # 統計情報
                    st.subheader("📊 出荷統計")
                    col_stat1, col_stat2, col_stat3 = st.columns(3)
                    
                    with col_stat1:
                        total_shipments = len(filtered_df)
                        st.metric("総出荷回数", f"{total_shipments}回")
                    
                    with col_stat2:
                        total_quantity = filtered_df['shipped_quantity'].sum()
                        st.metric("総出荷数量", f"{total_quantity:,.0f}個")
                    
                    with col_stat3:
                        unique_products = filtered_df['product_id'].nunique()
                        st.metric("出荷製品種類", f"{unique_products}種")
                else:
                    st.info("指定期間内の出荷実績がありません")
            else:
                st.info("出荷実績がありません")
        
        except Exception as e:
            st.error(f"出荷実績取得エラー: {e}")

    def _show_internal_orders(self):
        """社内注文（製造工程）タブ"""
        st.header("🏭 社内注文")
        st.write("積載計画で設定された数量を社内向けの加工指示として確認できます。")

        st.markdown("---")
        st.subheader("📅 表示期間")
        col1, col2 = st.columns(2)

        with col1:
            start_date = st.date_input(
                "開始日",
                value=date.today(),
                key="internal_order_start_date"
            )

        with col2:
            end_date = st.date_input(
                "終了日",
                value=date.today() + timedelta(days=7),
                key="internal_order_end_date"
            )

        if start_date > end_date:
            st.error("開始日は終了日以前の日付を選択してください。")
            return

        try:
            progress_df = self.service.get_delivery_progress(start_date, end_date)
        except Exception as e:
            st.error(f"データ取得エラー: {e}")
            return

        if progress_df is None or progress_df.empty:
            st.info("選択した期間に対象データがありません。")
            return

        if 'planned_quantity' not in progress_df.columns:
            st.warning("planned_quantity列が取得できませんでした。")
            return

        progress_df = progress_df[pd.to_numeric(progress_df['planned_quantity'], errors='coerce').fillna(0) > 0].copy()

        if progress_df.empty:
            st.info("積載計画の数量が設定されている製品がありません。")
            return

        progress_df['delivery_date'] = pd.to_datetime(progress_df['delivery_date']).dt.date


        st.subheader("📋 製品別マトリクス")
        matrix_df = self._create_internal_order_matrix(progress_df, start_date, end_date)
        if matrix_df.empty:
            st.info("表示対象データがありません。")
            return
        st.dataframe(
            matrix_df,
            use_container_width=True,
            hide_index=False,
            height=600
        )

        st.markdown("---")
        st.subheader("📥 Excel出力")
        excel_data = self._export_internal_orders_to_excel(matrix_df, start_date, end_date)
        filename = f"社内注文_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        st.download_button(
            label="📥 ダウンロード",
            data=excel_data,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="internal_order_excel_download"
        )

    def _create_internal_order_matrix(self, progress_df: pd.DataFrame, start_date: date, end_date: date) -> pd.DataFrame:
        """製品×納期のマトリクスを作成"""
        order_map = {}
        name_map = {}
        if hasattr(self.service, "product_repo"):
            try:
                master_df = self.service.product_repo.get_all_products()
            except Exception:
                master_df = pd.DataFrame()
            else:
                if isinstance(master_df, pd.DataFrame) and not master_df.empty:
                    temp_order = {}
                    temp_name = {}
                    for _, row in master_df.iterrows():
                        code = row.get('product_code')
                        if not code:
                            continue
                        display_val = row.get('display_id')
                        if pd.notna(display_val):
                            try:
                                display_val = int(display_val)
                            except (TypeError, ValueError):
                                pass
                        else:
                            display_val = None
                        temp_order[code] = display_val
                        temp_name[code] = row.get('product_name', '')
                    order_map = temp_order
                    name_map = temp_name

        product_codes = progress_df['product_code'].dropna().unique().tolist()

        def sort_key(code: str):
            display_value = order_map.get(code) if order_map else None
            if display_value is None or pd.isna(display_value):
                display_value = float('inf')
            return (display_value, code)

        product_codes.sort(key=sort_key)
        date_range = pd.date_range(start=start_date, end=end_date, freq='D')
        date_columns = [d.strftime('%Y/%m/%d') for d in date_range]
        date_values = [d.date() for d in date_range]

        matrix_data = []
        for product_code in product_codes:
            product_data = progress_df[progress_df['product_code'] == product_code]
            if product_data.empty:
                continue

            product_name = name_map.get(product_code)
            if not product_name:
                product_name = product_data['product_name'].iloc[0] if 'product_name' in product_data.columns else ''
            row = {
                '製品コード': product_code,
                '製品名': product_name
            }

            for date_obj, date_str in zip(date_values, date_columns):
                day_data = product_data[product_data['delivery_date'] == date_obj]
                if not day_data.empty:
                    planned_qty = pd.to_numeric(day_data['planned_quantity'], errors='coerce').fillna(0).sum()
                    row[date_str] = int(planned_qty) if planned_qty > 0 else 0
                else:
                    row[date_str] = 0

            matrix_data.append(row)

        matrix_df = pd.DataFrame(matrix_data)
        if not matrix_df.empty:
            matrix_df = matrix_df.set_index('製品コード')

        return matrix_df

    def _export_internal_orders_to_excel(self, matrix_df: pd.DataFrame, start_date: date, end_date: date):
        """マトリクスデータをExcelに出力"""
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            matrix_df.to_excel(writer, sheet_name='社内注文', index=True)

            workbook = writer.book
            worksheet = writer.sheets['社内注文']

            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            center_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                           min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.alignment = center_alignment
                    cell.border = border
                    if isinstance(cell.value, (int, float)) and cell.column > 2:
                        cell.number_format = '#,##0'

            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 30
            for col_idx in range(3, worksheet.max_column + 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = 12

            worksheet.insert_rows(1)
            worksheet['A1'] = f"社内注文 マトリクス（{start_date.strftime('%Y/%m/%d')} ～ {end_date.strftime('%Y/%m/%d')}）"
            worksheet['A1'].font = Font(bold=True, size=14)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=worksheet.max_column)
            worksheet['A1'].alignment = center_alignment

        output.seek(0)
        return output
